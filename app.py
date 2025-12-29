from __future__ import annotations

from pathlib import Path
import os, time, json, base64, hmac, hashlib, re
from io import BytesIO
from dataclasses import dataclass
from typing import Dict, Any, List, Tuple, Optional

import openpyxl
from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, Response
from jinja2 import Template


# ==========================================================
# PATHS + CONFIG
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
ROOT = BASE_DIR / "ROOT"
ROOT.mkdir(exist_ok=True)

SCALE_XLSX = ROOT / "3. SCM-Anwendungen(MA)_Gesamtbewertung.xlsx"
SUBMISSIONS_DIR = ROOT / "Form_Submissions"
SUBMISSIONS_DIR.mkdir(exist_ok=True)

APP_SECRET = os.getenv("SCM_FORM_SECRET", "CHANGE_ME_SECRET")
BASE_URL = os.getenv("SCM_FORM_BASE_URL", "http://localhost:8000")
TOKEN_TTL = int(os.getenv("SCM_TOKEN_TTL_SECONDS", str(7 * 24 * 3600)))
SCALE_SHEET = os.getenv("SCM_SCALE_SHEET", "Skala")

SOURCE_COL_NAME = "Datenherkunft / Bewertung"
SOURCE_MANUAL_VALUE = "Manuelle Bewertung"

CACHE_TTL = 30

app = FastAPI(title="SCM Form Backend")


# ==========================================================
# SMALL HELPERS
# ==========================================================
def norm(s: Any) -> str:
    s = "" if s is None else str(s)
    s = s.replace("\u00A0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def clean(s: Any) -> str:
    return "" if s is None else str(s).strip()

def sanitize_filename_part(s: Any) -> str:
    return re.sub(r"[^\w\-]", "_", clean(s))

def submission_path(round_id: str, supplier_id: str) -> Path:
    return SUBMISSIONS_DIR / f"submission_{sanitize_filename_part(supplier_id)}_R{sanitize_filename_part(round_id)}.json"


# ==========================================================
# TOKEN (HMAC)
# ==========================================================
def _b64e(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode("utf-8").rstrip("=")

def _b64d(s: str) -> bytes:
    return base64.urlsafe_b64decode(s + "==")

def make_token(payload: Dict[str, Any]) -> str:
    raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    sig = hmac.new(APP_SECRET.encode("utf-8"), raw, hashlib.sha256).digest()
    return f"{_b64e(raw)}.{_b64e(sig)}"

def read_token(token: str) -> Dict[str, Any]:
    try:
        part_raw, part_sig = token.split(".", 1)
        raw = _b64d(part_raw)
        sig = _b64d(part_sig)

        exp_sig = hmac.new(APP_SECRET.encode("utf-8"), raw, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, exp_sig):
            raise ValueError("bad signature")

        payload = json.loads(raw.decode("utf-8"))
        exp = payload.get("exp")
        if exp is not None and time.time() > float(exp):
            raise ValueError("expired")

        if not payload.get("supplier_id") or not payload.get("round_id"):
            raise ValueError("missing supplier_id/round_id")

        return payload
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid token: {e}")


# ==========================================================
# SCALE + MANUAL CRITERIA (single pass openpyxl)
# ==========================================================
def _find_header(ws, header_contains: str) -> Tuple[int, int]:
    tgt = norm(header_contains)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            if tgt in norm(cell.value):
                return cell.row, cell.column
    raise HTTPException(500, f"Header '{header_contains}' nicht gefunden (Sheet '{ws.title}').")

def _find_col_in_row(ws, row: int, needle_contains: str) -> Optional[int]:
    tgt = norm(needle_contains)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v is not None and tgt in norm(v):
            return c
    return None

def load_scale_and_manual() -> Tuple[Dict[str, Dict[int, str]], List[str]]:
    if not SCALE_XLSX.exists():
        raise HTTPException(500, f"Excel nicht gefunden: {SCALE_XLSX}")

    wb = openpyxl.load_workbook(SCALE_XLSX, data_only=True)
    if SCALE_SHEET not in wb.sheetnames:
        raise HTTPException(500, f"Sheet '{SCALE_SHEET}' nicht gefunden in Excel.")

    ws = wb[SCALE_SHEET]

    header_row, src_col = _find_header(ws, SOURCE_COL_NAME)
    crit_col = _find_col_in_row(ws, header_row, "kriter") or 1

    # Annahme wie bei dir: Skala-Spalten 0..100 sitzen ab Spalte 5 (E) bis 10 (J)
    # (Weil du vorher df.iloc[i,4..9] genutzt hast)
    scale_cols = {0: 5, 20: 6, 40: 7, 60: 8, 80: 9, 100: 10}

    scale: Dict[str, Dict[int, str]] = {}
    manual: List[str] = []

    manual_marker = norm(SOURCE_MANUAL_VALUE)

    for r in range(header_row + 1, ws.max_row + 1):
        crit = clean(ws.cell(r, crit_col).value)
        if not crit or norm(crit) in ("nan", "none"):
            continue

        # Skala sammeln
        scale[crit] = {pts: clean(ws.cell(r, col).value) for pts, col in scale_cols.items()}

        # Manual markieren
        src_val = ws.cell(r, src_col).value
        if norm(src_val) == manual_marker:
            manual.append(crit)

    if not scale:
        raise HTTPException(500, "Skala konnte nicht geladen werden (keine Kriterien gefunden).")

    if not manual:
        raise HTTPException(
            500,
            f"Header gefunden (Row {header_row}, Col {src_col}), aber keine Zeilen mit '{SOURCE_MANUAL_VALUE}'."
        )

    # unique, Reihenfolge wie Skala
    manual = list(dict.fromkeys(manual))
    return scale, manual


@dataclass
class _Cache:
    ts: float = 0.0
    scale: Optional[Dict[str, Dict[int, str]]] = None
    manual: Optional[List[str]] = None

_CACHE = _Cache()

def get_scale_and_manual() -> Tuple[Dict[str, Dict[int, str]], List[str]]:
    now = time.time()
    if _CACHE.scale is not None and (now - _CACHE.ts) < CACHE_TTL:
        return _CACHE.scale, _CACHE.manual or []
    scale, manual = load_scale_and_manual()
    _CACHE.ts, _CACHE.scale, _CACHE.manual = now, scale, manual
    return scale, manual


# ==========================================================
# XLSX Builder (für Bot)
# ==========================================================
def build_reply_xlsx(answers: Dict[str, int]) -> bytes:
    # minimal: openpyxl direkt (spart pandas)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antwort"
    ws.append(["Kriterium", "Bewertung"])
    for k, v in answers.items():
        ws.append([k, int(v)])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ==========================================================
# HTML Template
# ==========================================================
FORM_HTML = Template("""
<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>SCM Bewertung</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 24px; max-width: 980px; }
    .meta { color:#555; margin-bottom: 16px; }
    .card { border:1px solid #ddd; border-radius: 10px; padding: 14px; margin: 12px 0; }
    .crit { font-weight: 700; margin-bottom: 8px; }
    .opt { margin: 8px 0; }
    .desc { color:#444; font-size: 13px; margin-left: 22px; white-space: pre-wrap; }
    .btn { padding: 10px 14px; border-radius: 10px; border:0; background:#1a73e8; color:white; font-weight:700; cursor:pointer; }
    .subtle { color:#666; font-size: 12px; }
  </style>
</head>
<body>
  <h2>SCM Bewertung</h2>
  <div class="meta">
    Runde: <b>{{ round_id }}</b> • Lieferant: <b>{{ supplier_id }}</b>
  </div>

  <form method="post" action="/submit">
    <input type="hidden" name="token" value="{{ token }}"/>

    {% for i, item in items %}
      <div class="card">
        <div class="crit">{{ item.crit }}</div>
        {% for pts in [100,80,60,40,20,0] %}
          <div class="opt">
            <label>
              <input type="radio" name="c_{{ i }}" value="{{ pts }}" required />
              <b>{{ pts }}</b> Punkte
            </label>
            {% set d = item.scale.get(pts, "") %}
            {% if d %}
              <div class="desc">{{ d }}</div>
            {% endif %}
          </div>
        {% endfor %}
      </div>
    {% endfor %}

    <button class="btn" type="submit">Absenden</button>
    <div class="subtle" style="margin-top:10px;">
      Nach dem Absenden werden die Daten serverseitig gespeichert. Der Bot holt sie automatisch ab.
    </div>
  </form>
</body>
</html>
""")


# ==========================================================
# Storage
# ==========================================================
def save_submission(round_id: str, supplier_id: str, answers: Dict[str, int]) -> Dict[str, Any]:
    data = {"round_id": round_id, "supplier_id": supplier_id, "submitted_at": time.time(), "answers": answers}
    submission_path(round_id, supplier_id).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data

def load_submission(round_id: str, supplier_id: str) -> Dict[str, Any]:
    p = submission_path(round_id, supplier_id)
    if not p.exists():
        raise HTTPException(404, "submission not found")
    return json.loads(p.read_text(encoding="utf-8"))


# ==========================================================
# Routes
# ==========================================================
@app.get("/", response_class=HTMLResponse)
async def home():
    scale, manual = get_scale_and_manual()
    return HTMLResponse(
        f"""
        <h3>SCM Formular-Service ✅</h3>
        <p>Excel: <code>{SCALE_XLSX}</code></p>
        <p>Manuelle Kriterien (nach '{SOURCE_COL_NAME}' = '{SOURCE_MANUAL_VALUE}'): <b>{len(manual)}</b></p>
        <p>Test: <code>{BASE_URL}/issue-link?supplier_id=K_1&round_id=12345678</code></p>
        """
    )

@app.get("/issue-link")
async def issue_link(supplier_id: str, round_id: str):
    supplier_id, round_id = clean(supplier_id), clean(round_id)
    if not supplier_id or not round_id:
        raise HTTPException(400, "missing supplier_id or round_id")

    token = make_token({"supplier_id": supplier_id, "round_id": round_id, "exp": time.time() + TOKEN_TTL})
    url = f"{BASE_URL.rstrip('/')}/evaluate?token={token}"
    return JSONResponse({"url": url, "token": token})

@app.get("/evaluate", response_class=HTMLResponse)
async def evaluate(token: str):
    payload = read_token(token)
    scale, manual = get_scale_and_manual()

    items = [{"crit": c, "scale": scale[c]} for c in manual]
    html = FORM_HTML.render(
        token=token,
        round_id=payload["round_id"],
        supplier_id=payload["supplier_id"],
        items=list(enumerate(items)),
    )
    return HTMLResponse(html)

@app.post("/submit", response_class=HTMLResponse)
async def submit(request: Request, token: str = Form(...)):
    payload = read_token(token)
    scale, manual = get_scale_and_manual()

    form = await request.form()
    answers: Dict[str, int] = {}

    for i, crit in enumerate(manual):
        val = form.get(f"c_{i}")
        if val is None:
            raise HTTPException(400, f"missing answer for {crit}")
        pts = int(val)
        if pts not in (0, 20, 40, 60, 80, 100):
            raise HTTPException(400, f"invalid scale value: {pts}")
        answers[crit] = pts

    save_submission(payload["round_id"], payload["supplier_id"], answers)
    return HTMLResponse(
        f"""
        <h3>Danke! ✅</h3>
        <p>Ihre Bewertung für <b>{payload["supplier_id"]}</b> (Runde <b>{payload["round_id"]}</b>) wurde gespeichert.</p>
        <p>Sie können dieses Fenster jetzt schließen.</p>
        """
    )

@app.get("/api/submissions")
async def api_submissions(round_id: str):
    rid = clean(round_id)
    out = []
    for f in SUBMISSIONS_DIR.glob("submission_*_R*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            if str(data.get("round_id")) == rid:
                out.append({"supplier_id": data.get("supplier_id"), "submitted_at": float(data.get("submitted_at", 0))})
        except Exception:
            continue
    out.sort(key=lambda x: x["submitted_at"])
    return JSONResponse(out)

@app.get("/api/xlsx")
async def api_xlsx(round_id: str, supplier_id: str):
    data = load_submission(round_id, supplier_id)
    answers = {k: int(v) for k, v in (data.get("answers") or {}).items()}
    xlsx_bytes = build_reply_xlsx(answers)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Antwort_{supplier_id}_R{round_id}.xlsx"'},
    )
